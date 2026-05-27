"""
Professional financial CSV reporting template.

Quick run from the Coffee_Sales folder:
    python financial_csv_report_template.py --input raw_coffee_sales.csv

Default outputs:
    financial_report_output/<file_name>_cleaned.csv
    financial_report_output/<file_name>_finance_fact.csv
    financial_report_output/<file_name>_financial_report.md
    financial_report_output/<file_name>_financial_workbook.xlsx
    financial_report_output/charts/*.png

This template is designed for reusable finance reporting work:
- Cleans common CSV issues: column names, blank strings, currencies, dates,
  percentages, duplicate rows, and missing values.
- Detects common financial structures: amount, debit/credit, revenue/cost,
  Actual/Budget/Forecast/Prior columns, account, account type, scenario, and
  dimensions such as department, entity, region, store, product, or cost center.
- Builds a normalized finance fact table with P&L sign convention:
  income positive, costs and expenses negative.
- Generates P&L summaries, monthly trends, optional variance analysis, charts,
  a Markdown report, and an Excel workbook for audit-friendly review.
"""

from __future__ import annotations

import argparse
import re
import textwrap
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Iterable

import matplotlib

matplotlib.use("Agg")

import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
from pandas.api.types import (
    is_datetime64_any_dtype,
    is_numeric_dtype,
    is_object_dtype,
    is_string_dtype,
)


DATE_KEYWORDS = ("date", "day", "month", "period", "time", "posted", "created", "updated")
ID_KEYWORDS = ("id", "key", "code", "uuid", "number", "no")
AMOUNT_KEYWORDS = (
    "amount",
    "value",
    "balance",
    "revenue",
    "sales",
    "income",
    "cost",
    "expense",
    "debit",
    "credit",
    "discount",
    "percent",
    "rate",
    "actual",
    "budget",
    "forecast",
    "prior",
    "py",
    "ly",
)
NON_DIMENSION_MEASURE_KEYWORDS = ("unit", "units", "quantity", "qty", "price", "discount", "percent", "rate")
DIMENSION_PRIORITY = (
    "entity",
    "company",
    "business_unit",
    "division",
    "department",
    "cost_center",
    "location",
    "region",
    "store",
    "branch",
    "channel",
    "customer",
    "vendor",
    "supplier",
    "product",
    "category",
    "project",
    "manager",
)
TIME_FEATURES = {"period", "month", "quarter", "fiscal_year", "fiscal_quarter", "year"}
P_AND_L_ORDER = (
    "Revenue",
    "COGS",
    "Gross Profit",
    "Operating Expense",
    "Operating Profit",
    "Other Income",
    "Other Expense",
    "Tax",
    "Net Profit",
    "Unclassified Income",
    "Unclassified Expense",
    "Unclassified",
)


@dataclass
class FinancialRoles:
    date: str | None = None
    amount: str | None = None
    debit: str | None = None
    credit: str | None = None
    revenue: str | None = None
    cost: str | None = None
    actual: str | None = None
    budget: str | None = None
    forecast: str | None = None
    prior: str | None = None
    account: str | None = None
    account_type: str | None = None
    line_item: str | None = None
    scenario: str | None = None
    dimensions: list[str] | None = None
    numeric: list[str] | None = None

    def __post_init__(self) -> None:
        self.dimensions = self.dimensions or []
        self.numeric = self.numeric or []


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Clean financial CSVs and generate a professional finance report.",
        formatter_class=argparse.ArgumentDefaultsHelpFormatter,
    )
    parser.add_argument("--input", default="raw_coffee_sales.csv", help="Path to the CSV file to analyze.")
    parser.add_argument("--output-dir", default="financial_report_output", help="Output folder.")
    parser.add_argument("--encoding", default="utf-8-sig", help="CSV encoding; falls back to cp1252 if needed.")
    parser.add_argument("--top-n", type=int, default=10, help="Number of top drivers to show.")
    parser.add_argument("--currency-symbol", default="$", help="Currency symbol used in the report.")
    parser.add_argument(
        "--sign-mode",
        choices=("income-positive", "as-is"),
        default="income-positive",
        help="Normalize P&L signs or keep source signs as-is.",
    )
    parser.add_argument(
        "--debit-credit-convention",
        choices=("debit-minus-credit", "credit-minus-debit"),
        default="debit-minus-credit",
        help="How to compute amount when only debit and credit columns exist.",
    )
    parser.add_argument("--fiscal-year-start", type=int, default=1, help="Fiscal year start month, 1-12.")
    parser.add_argument("--no-title-case", action="store_true", help="Do not auto title-case dimension values.")

    parser.add_argument("--date-column", help="Override detected date column.")
    parser.add_argument("--amount-column", help="Override detected amount column.")
    parser.add_argument("--debit-column", help="Override detected debit column.")
    parser.add_argument("--credit-column", help="Override detected credit column.")
    parser.add_argument("--revenue-column", help="Override detected revenue column.")
    parser.add_argument("--cost-column", help="Override detected cost/COGS column.")
    parser.add_argument("--actual-column", help="Override detected Actual column.")
    parser.add_argument("--budget-column", help="Override detected Budget column.")
    parser.add_argument("--forecast-column", help="Override detected Forecast column.")
    parser.add_argument("--prior-column", help="Override detected Prior Year/Prior Period column.")
    parser.add_argument("--account-column", help="Override detected account column.")
    parser.add_argument("--account-type-column", help="Override detected account type column.")
    parser.add_argument("--line-item-column", help="Override detected financial line item column.")
    parser.add_argument("--scenario-column", help="Override detected scenario/version column.")
    return parser.parse_args()


def snake_case(name: object) -> str:
    text = str(name).strip()
    text = re.sub(r"(?<=[a-z0-9])(?=[A-Z])", "_", text)
    text = re.sub(r"[^0-9a-zA-Z]+", "_", text)
    text = text.strip("_").lower()
    text = re.sub(r"_+", "_", text)
    return text or "unnamed_column"


def make_unique(names: Iterable[str]) -> list[str]:
    seen: dict[str, int] = {}
    unique_names: list[str] = []
    for name in names:
        count = seen.get(name, 0)
        unique_name = name if count == 0 else f"{name}_{count + 1}"
        seen[name] = count + 1
        unique_names.append(unique_name)
    return unique_names


def normalize_override(value: str | None) -> str | None:
    return snake_case(value) if value else None


def require_override(columns: Iterable[str], value: str | None, label: str) -> str | None:
    normalized = normalize_override(value)
    if normalized and normalized not in set(columns):
        raise ValueError(f"{label} override not found after column normalization: {value} -> {normalized}")
    return normalized


def is_id_like(column: str) -> bool:
    tokens = set(column.split("_"))
    return any(keyword in tokens or column.endswith(f"_{keyword}") for keyword in ID_KEYWORDS)


def is_date_like_name(column: str) -> bool:
    return any(keyword in column for keyword in DATE_KEYWORDS)


def is_amount_like_name(column: str) -> bool:
    return any(keyword in column for keyword in AMOUNT_KEYWORDS)


def is_dimension_candidate(df: pd.DataFrame, column: str) -> bool:
    if column in TIME_FEATURES or is_id_like(column) or is_date_like_name(column):
        return False
    if is_amount_like_name(column) or any(keyword in column for keyword in NON_DIMENSION_MEASURE_KEYWORDS):
        return False
    if is_object_dtype(df[column]) or is_string_dtype(df[column]):
        return True
    if is_numeric_dtype(df[column]):
        non_null = int(df[column].notna().sum())
        if non_null == 0:
            return False
        unique_count = int(df[column].nunique(dropna=True))
        return unique_count <= min(12, max(3, non_null // 10))
    return False


def clean_string_values(series: pd.Series) -> pd.Series:
    cleaned = series.astype("string").str.strip()
    cleaned = cleaned.str.replace(r"\s+", " ", regex=True)
    blank_tokens = {"", "na", "n/a", "nan", "none", "null", "-", "--", "not available"}
    return cleaned.mask(cleaned.str.lower().isin(blank_tokens))


def parse_date_series(series: pd.Series) -> tuple[pd.Series, float]:
    non_null = series.dropna()
    if non_null.empty:
        return pd.to_datetime(series, errors="coerce"), 0.0
    try:
        parsed = pd.to_datetime(series, errors="coerce", format="mixed")
    except TypeError:
        parsed = pd.to_datetime(series, errors="coerce")
    ratio = float(parsed.notna().sum() / max(non_null.shape[0], 1))
    return parsed, ratio


def parse_numeric_series(series: pd.Series) -> tuple[pd.Series, float]:
    non_null = series.dropna()
    if non_null.empty:
        return pd.to_numeric(series, errors="coerce"), 0.0
    cleaned = series.astype("string").str.strip()
    cleaned = cleaned.str.replace(r"^\((.*)\)$", r"-\1", regex=True)
    cleaned = cleaned.str.replace(",", "", regex=False)
    cleaned = cleaned.str.replace("$", "", regex=False)
    cleaned = cleaned.str.replace("%", "", regex=False)
    cleaned = cleaned.str.replace(r"[^0-9.\-]", "", regex=True)
    numeric = pd.to_numeric(cleaned, errors="coerce")
    ratio = float(numeric.notna().sum() / max(non_null.shape[0], 1))
    return numeric, ratio


def smart_title(value: object) -> object:
    if pd.isna(value):
        return value
    text = str(value).strip()
    if not text:
        return pd.NA
    if re.fullmatch(r"[A-Z0-9_-]{2,}", text):
        return text
    if "@" in text or text.startswith(("http://", "https://")):
        return text
    return text.title()


def should_title_case(series: pd.Series, column: str) -> bool:
    if is_id_like(column) or is_amount_like_name(column) or is_date_like_name(column):
        return False
    non_null_count = int(series.notna().sum())
    if non_null_count == 0:
        return False
    unique_ratio = series.nunique(dropna=True) / max(non_null_count, 1)
    return not (non_null_count > 20 and unique_ratio > 0.75)


def find_first_column(
    columns: list[str],
    exact: Iterable[str],
    contains: Iterable[str],
    exclude: Iterable[str] = (),
) -> str | None:
    exact_set = set(exact)
    excluded = tuple(exclude)
    for column in columns:
        if column in exact_set and not any(token in column for token in excluded):
            return column
    for column in columns:
        if any(token in column for token in contains) and not any(token in column for token in excluded):
            return column
    return None


def load_csv(path: Path, encoding: str) -> pd.DataFrame:
    try:
        return pd.read_csv(path, encoding=encoding)
    except UnicodeDecodeError:
        return pd.read_csv(path, encoding="cp1252")


def clean_dataframe(raw_df: pd.DataFrame, *, title_case_text: bool) -> tuple[pd.DataFrame, dict[str, object]]:
    cleaned = raw_df.copy()
    original_columns = list(cleaned.columns)
    normalized_columns = make_unique(snake_case(column) for column in cleaned.columns)
    cleaned.columns = normalized_columns

    profile: dict[str, object] = {
        "original_shape": raw_df.shape,
        "column_mapping": list(zip(original_columns, normalized_columns)),
        "parsed_date_columns": {},
        "converted_numeric_columns": {},
        "filled_missing_values": {},
        "duplicate_rows_removed": 0,
    }

    for column in cleaned.columns:
        if is_object_dtype(cleaned[column]) or is_string_dtype(cleaned[column]):
            cleaned[column] = clean_string_values(cleaned[column])

    profile["missing_before"] = cleaned.isna().sum().to_dict()

    parsed_dates: dict[str, float] = {}
    for column in list(cleaned.columns):
        if not (is_object_dtype(cleaned[column]) or is_string_dtype(cleaned[column])):
            continue
        if is_id_like(column):
            continue
        parsed, ratio = parse_date_series(cleaned[column])
        if (is_date_like_name(column) and ratio >= 0.50) or ratio >= 0.95:
            cleaned[column] = parsed
            parsed_dates[column] = round(ratio, 3)

    converted_numeric: dict[str, float] = {}
    for column in list(cleaned.columns):
        if is_datetime64_any_dtype(cleaned[column]) or is_id_like(column):
            continue
        if not (is_object_dtype(cleaned[column]) or is_string_dtype(cleaned[column])):
            continue
        numeric, ratio = parse_numeric_series(cleaned[column])
        if (is_amount_like_name(column) and ratio >= 0.50) or ratio >= 0.95:
            cleaned[column] = numeric
            converted_numeric[column] = round(ratio, 3)

    for column in list(cleaned.columns):
        if not (is_object_dtype(cleaned[column]) or is_string_dtype(cleaned[column])):
            continue
        cleaned[column] = clean_string_values(cleaned[column])
        if title_case_text and should_title_case(cleaned[column], column):
            cleaned[column] = cleaned[column].map(smart_title).astype("string")

    duplicate_rows = int(cleaned.duplicated().sum())
    if duplicate_rows:
        cleaned = cleaned.drop_duplicates().reset_index(drop=True)

    filled_missing: dict[str, str] = {}
    for column in cleaned.columns:
        missing_count = int(cleaned[column].isna().sum())
        if missing_count == 0:
            continue
        if is_datetime64_any_dtype(cleaned[column]):
            filled_missing[column] = "kept as missing date"
            continue
        if is_numeric_dtype(cleaned[column]):
            if is_amount_like_name(column):
                cleaned[column] = cleaned[column].fillna(0)
                filled_missing[column] = f"filled {missing_count} with 0"
            else:
                median = cleaned[column].median()
                fill_value = 0 if pd.isna(median) else median
                cleaned[column] = cleaned[column].fillna(fill_value)
                filled_missing[column] = f"filled {missing_count} with median {fill_value}"
            continue
        cleaned[column] = cleaned[column].fillna("Unassigned")
        filled_missing[column] = f"filled {missing_count} with Unassigned"

    profile["parsed_date_columns"] = parsed_dates
    profile["converted_numeric_columns"] = converted_numeric
    profile["duplicate_rows_removed"] = duplicate_rows
    profile["filled_missing_values"] = filled_missing
    profile["missing_after_cleaning"] = cleaned.isna().sum().to_dict()
    profile["cleaned_shape"] = cleaned.shape
    return cleaned, profile


def infer_roles(df: pd.DataFrame, args: argparse.Namespace) -> FinancialRoles:
    columns = list(df.columns)
    numeric_columns = [column for column in columns if is_numeric_dtype(df[column])]

    roles = FinancialRoles()
    roles.date = require_override(columns, args.date_column, "date-column") or find_first_column(
        columns,
        exact=("date", "transaction_date", "posting_date", "posted_date", "invoice_date", "order_date"),
        contains=("date", "posted", "posting"),
    )
    if roles.date is None:
        for column in columns:
            if is_datetime64_any_dtype(df[column]):
                roles.date = column
                break

    roles.amount = require_override(columns, args.amount_column, "amount-column") or find_first_column(
        numeric_columns,
        exact=("amount", "net_amount", "transaction_amount", "value", "balance"),
        contains=("amount", "value", "balance"),
        exclude=("discount", "tax_rate", "percent", "margin", "price", "unit"),
    )
    roles.debit = require_override(columns, args.debit_column, "debit-column") or find_first_column(
        numeric_columns,
        exact=("debit", "debit_amount"),
        contains=("debit",),
    )
    roles.credit = require_override(columns, args.credit_column, "credit-column") or find_first_column(
        numeric_columns,
        exact=("credit", "credit_amount"),
        contains=("credit",),
    )
    roles.revenue = require_override(columns, args.revenue_column, "revenue-column") or find_first_column(
        numeric_columns,
        exact=("revenue", "sales", "net_sales", "gross_sales", "income", "turnover"),
        contains=("revenue", "sales", "income", "turnover"),
        exclude=("cost", "expense", "discount", "unit"),
    )
    roles.cost = require_override(columns, args.cost_column, "cost-column") or find_first_column(
        numeric_columns,
        exact=("cost", "cogs", "cost_of_goods_sold", "cost_of_sales", "total_cost"),
        contains=("cost", "cogs"),
        exclude=("center", "customer"),
    )
    roles.actual = require_override(columns, args.actual_column, "actual-column") or find_first_column(
        numeric_columns,
        exact=("actual", "actuals", "actual_amount"),
        contains=("actual",),
    )
    roles.budget = require_override(columns, args.budget_column, "budget-column") or find_first_column(
        numeric_columns,
        exact=("budget", "budget_amount", "plan"),
        contains=("budget", "plan"),
    )
    roles.forecast = require_override(columns, args.forecast_column, "forecast-column") or find_first_column(
        numeric_columns,
        exact=("forecast", "forecast_amount", "fcst"),
        contains=("forecast", "fcst"),
    )
    roles.prior = require_override(columns, args.prior_column, "prior-column") or find_first_column(
        numeric_columns,
        exact=("prior", "prior_year", "prior_period", "py", "ly", "last_year"),
        contains=("prior", "last_year"),
    )
    roles.account = require_override(columns, args.account_column, "account-column") or find_first_column(
        columns,
        exact=("account", "account_name", "gl_account", "ledger_account"),
        contains=("account", "gl_"),
        exclude=("account_id", "account_code", "account_number"),
    )
    roles.account_type = require_override(columns, args.account_type_column, "account-type-column") or find_first_column(
        columns,
        exact=("account_type", "account_class", "classification", "statement_type"),
        contains=("account_type", "classification"),
    )
    roles.line_item = require_override(columns, args.line_item_column, "line-item-column") or find_first_column(
        columns,
        exact=("line_item", "financial_line", "statement_line", "pl_line", "p_l_line", "pnl_line"),
        contains=("line_item", "financial_line", "statement_line", "pl_line", "pnl_line"),
    )
    roles.scenario = require_override(columns, args.scenario_column, "scenario-column") or find_first_column(
        columns,
        exact=("scenario", "version", "case", "data_type"),
        contains=("scenario", "version"),
    )
    roles.numeric = numeric_columns
    roles.dimensions = infer_dimensions(df, roles)
    return roles


def infer_dimensions(df: pd.DataFrame, roles: FinancialRoles) -> list[str]:
    excluded = {
        roles.date,
        roles.amount,
        roles.debit,
        roles.credit,
        roles.revenue,
        roles.cost,
        roles.actual,
        roles.budget,
        roles.forecast,
        roles.prior,
        roles.account,
        roles.account_type,
        roles.line_item,
        roles.scenario,
        None,
    }
    dimensions: list[str] = []
    for keyword in DIMENSION_PRIORITY:
        for column in df.columns:
            if column in excluded or column in dimensions:
                continue
            if keyword in column and is_dimension_candidate(df, column):
                dimensions.append(column)

    for column in df.columns:
        if column in excluded or column in dimensions:
            continue
        if is_dimension_candidate(df, column):
            dimensions.append(column)
    return dimensions[:10]


def contains_any(value: object, keywords: Iterable[str]) -> bool:
    if pd.isna(value):
        return False
    text = str(value).strip().lower()
    return any(keyword in text for keyword in keywords)


def classify_financial_line(
    *,
    line_item: object = None,
    account_type: object = None,
    account: object = None,
    amount: float | None = None,
) -> str:
    combined = " ".join(str(value).lower() for value in (line_item, account_type, account) if not pd.isna(value))
    combined = combined.replace("&", "and")

    if any(token in combined for token in ("revenue", "sales", "income", "turnover")):
        if any(token in combined for token in ("interest expense", "tax expense")):
            return "Other Expense"
        if "other income" in combined or "interest income" in combined:
            return "Other Income"
        return "Revenue"
    if any(token in combined for token in ("cogs", "cost of goods", "cost of sales", "direct cost", "cost of revenue")):
        return "COGS"
    if any(token in combined for token in ("tax", "income tax")):
        return "Tax"
    if any(token in combined for token in ("interest expense", "financing cost", "bank fee", "other expense")):
        return "Other Expense"
    if any(token in combined for token in ("opex", "operating expense", "sg&a", "sga", "payroll", "salary", "rent")):
        return "Operating Expense"
    if "expense" in combined or "cost" in combined:
        return "Operating Expense"

    if amount is not None and not pd.isna(amount):
        if amount > 0:
            return "Unclassified Income"
        if amount < 0:
            return "Unclassified Expense"
    return "Unclassified"


def normalize_pnl_amount(amount: object, line: str, sign_mode: str) -> float:
    if pd.isna(amount):
        return 0.0
    value = float(amount)
    if sign_mode == "as-is":
        return value
    if line in {"Revenue", "Other Income", "Unclassified Income"}:
        return abs(value)
    if line in {"COGS", "Operating Expense", "Other Expense", "Tax", "Unclassified Expense"}:
        return -abs(value)
    return value


def build_base_fact_columns(df: pd.DataFrame, roles: FinancialRoles) -> pd.DataFrame:
    fact = pd.DataFrame({"source_row_id": df.index + 1})
    if roles.date:
        fact["transaction_date"] = df[roles.date]
    else:
        fact["transaction_date"] = pd.NaT
    if roles.account:
        fact["account"] = df[roles.account].astype("string")
    else:
        fact["account"] = "Unassigned"
    if roles.account_type:
        fact["account_type"] = df[roles.account_type].astype("string")
    else:
        fact["account_type"] = "Unassigned"
    if roles.scenario:
        fact["scenario"] = df[roles.scenario].astype("string")
    else:
        fact["scenario"] = "Actual"
    for dimension in roles.dimensions:
        fact[dimension] = df[dimension]
    return fact


def add_time_fields(fact: pd.DataFrame, fiscal_year_start: int) -> pd.DataFrame:
    enriched = fact.copy()
    if "transaction_date" in enriched and is_datetime64_any_dtype(enriched["transaction_date"]):
        dates = enriched["transaction_date"]
        enriched["period"] = dates.dt.to_period("M").astype("string").fillna("Unassigned Period")
        fiscal_month = ((dates.dt.month - fiscal_year_start) % 12) + 1
        fiscal_year = dates.dt.year + (dates.dt.month >= fiscal_year_start).astype(int)
        if fiscal_year_start == 1:
            fiscal_year = dates.dt.year
        enriched["fiscal_year"] = fiscal_year.astype("Int64")
        enriched["fiscal_quarter"] = ("Q" + (((fiscal_month - 1) // 3) + 1).astype("Int64").astype("string")).fillna("Unassigned")
    else:
        enriched["period"] = "Unassigned Period"
        enriched["fiscal_year"] = pd.NA
        enriched["fiscal_quarter"] = "Unassigned"
    return enriched


def build_finance_fact(df: pd.DataFrame, roles: FinancialRoles, args: argparse.Namespace) -> tuple[pd.DataFrame, dict[str, object]]:
    source_info: dict[str, object] = {"amount_source": "not detected", "generated_rows": 0}
    frames: list[pd.DataFrame] = []
    base = build_base_fact_columns(df, roles)

    scenario_measure_columns = [
        ("Actual", roles.actual),
        ("Budget", roles.budget),
        ("Forecast", roles.forecast),
        ("Prior", roles.prior),
    ]
    available_scenario_measures = [(name, column) for name, column in scenario_measure_columns if column]

    if available_scenario_measures and not roles.amount and not (roles.debit and roles.credit):
        for scenario_name, column in available_scenario_measures:
            scenario_frame = base.copy()
            scenario_frame["scenario"] = scenario_name
            scenario_frame["source_amount"] = df[column]
            scenario_frame["financial_line"] = [
                classify_financial_line(
                    line_item=df.at[index, roles.line_item] if roles.line_item else None,
                    account_type=df.at[index, roles.account_type] if roles.account_type else None,
                    account=df.at[index, roles.account] if roles.account else None,
                    amount=df.at[index, column],
                )
                for index in df.index
            ]
            scenario_frame["amount"] = [
                normalize_pnl_amount(amount, line, args.sign_mode)
                for amount, line in zip(scenario_frame["source_amount"], scenario_frame["financial_line"])
            ]
            frames.append(scenario_frame)
        source_info["amount_source"] = "scenario measure columns"

    elif roles.amount or (roles.debit and roles.credit):
        amount_frame = base.copy()
        if roles.amount:
            amount_frame["source_amount"] = df[roles.amount]
            source_info["amount_source"] = roles.amount
        else:
            debit = df[roles.debit].fillna(0)
            credit = df[roles.credit].fillna(0)
            if args.debit_credit_convention == "credit-minus-debit":
                amount_frame["source_amount"] = credit - debit
            else:
                amount_frame["source_amount"] = debit - credit
            source_info["amount_source"] = f"{roles.debit} and {roles.credit}"
        amount_frame["financial_line"] = [
            classify_financial_line(
                line_item=df.at[index, roles.line_item] if roles.line_item else None,
                account_type=df.at[index, roles.account_type] if roles.account_type else None,
                account=df.at[index, roles.account] if roles.account else None,
                amount=amount_frame.at[position, "source_amount"],
            )
            for position, index in enumerate(df.index)
        ]
        amount_frame["amount"] = [
            normalize_pnl_amount(amount, line, args.sign_mode)
            for amount, line in zip(amount_frame["source_amount"], amount_frame["financial_line"])
        ]
        frames.append(amount_frame)

    elif roles.revenue or roles.cost:
        if roles.revenue:
            revenue_frame = base.copy()
            revenue_frame["financial_line"] = "Revenue"
            revenue_frame["source_amount"] = df[roles.revenue]
            revenue_frame["amount"] = [
                normalize_pnl_amount(amount, "Revenue", args.sign_mode) for amount in revenue_frame["source_amount"]
            ]
            frames.append(revenue_frame)
        if roles.cost:
            cost_frame = base.copy()
            cost_frame["financial_line"] = "COGS"
            cost_frame["source_amount"] = df[roles.cost]
            cost_frame["amount"] = [
                normalize_pnl_amount(amount, "COGS", args.sign_mode) for amount in cost_frame["source_amount"]
            ]
            frames.append(cost_frame)
        source_info["amount_source"] = "revenue/cost columns"

    if not frames:
        raise ValueError(
            "No financial amount source found. Provide an amount column, debit/credit columns, "
            "Actual/Budget style columns, or revenue/cost columns."
        )

    fact = pd.concat(frames, ignore_index=True)
    fact["source_amount"] = pd.to_numeric(fact["source_amount"], errors="coerce").fillna(0.0)
    fact["amount"] = pd.to_numeric(fact["amount"], errors="coerce").fillna(0.0)
    fact["abs_amount"] = fact["amount"].abs()
    fact = add_time_fields(fact, args.fiscal_year_start)
    source_info["generated_rows"] = len(fact)
    source_info["line_items"] = fact["financial_line"].value_counts().to_dict()
    return fact, source_info


def select_actual_fact(fact: pd.DataFrame) -> pd.DataFrame:
    if "scenario" not in fact.columns:
        return fact
    scenario_values = fact["scenario"].astype("string").str.lower()
    if scenario_values.eq("actual").any():
        return fact[scenario_values.eq("actual")].copy()
    return fact.copy()


def pnl_metrics(fact: pd.DataFrame) -> dict[str, float]:
    actual_fact = select_actual_fact(fact)
    totals = actual_fact.groupby("financial_line")["amount"].sum().to_dict()
    revenue = totals.get("Revenue", 0.0) + totals.get("Unclassified Income", 0.0)
    cogs = totals.get("COGS", 0.0)
    gross_profit = revenue + cogs
    opex = totals.get("Operating Expense", 0.0)
    operating_profit = gross_profit + opex
    other_income = totals.get("Other Income", 0.0)
    other_expense = totals.get("Other Expense", 0.0)
    tax = totals.get("Tax", 0.0)
    unclassified = totals.get("Unclassified", 0.0) + totals.get("Unclassified Expense", 0.0)
    net_profit = operating_profit + other_income + other_expense + tax + unclassified
    total_expenses = abs(cogs) + abs(opex) + abs(other_expense) + abs(tax) + abs(totals.get("Unclassified Expense", 0.0))
    return {
        "Revenue": revenue,
        "COGS": cogs,
        "Gross Profit": gross_profit,
        "Operating Expense": opex,
        "Operating Profit": operating_profit,
        "Other Income": other_income,
        "Other Expense": other_expense,
        "Tax": tax,
        "Unclassified": unclassified,
        "Net Profit": net_profit,
        "Total Expenses": total_expenses,
        "Gross Margin": gross_profit / revenue if revenue else np.nan,
        "Operating Margin": operating_profit / revenue if revenue else np.nan,
        "Net Margin": net_profit / revenue if revenue else np.nan,
    }


def build_pnl_summary(fact: pd.DataFrame) -> pd.DataFrame:
    metrics = pnl_metrics(fact)
    rows = [
        ("Revenue", metrics["Revenue"]),
        ("COGS", metrics["COGS"]),
        ("Gross Profit", metrics["Gross Profit"]),
        ("Operating Expense", metrics["Operating Expense"]),
        ("Operating Profit", metrics["Operating Profit"]),
        ("Other Income", metrics["Other Income"]),
        ("Other Expense", metrics["Other Expense"]),
        ("Tax", metrics["Tax"]),
        ("Unclassified", metrics["Unclassified"]),
        ("Net Profit", metrics["Net Profit"]),
    ]
    revenue = metrics["Revenue"]
    summary = pd.DataFrame(rows, columns=["line_item", "amount"])
    summary["percent_of_revenue"] = summary["amount"] / revenue if revenue else np.nan
    summary = summary[(summary["amount"].abs() > 0.000001) | summary["line_item"].isin(["Revenue", "Net Profit"])]
    summary["sort_order"] = summary["line_item"].map({name: index for index, name in enumerate(P_AND_L_ORDER)}).fillna(99)
    return summary.sort_values("sort_order").drop(columns="sort_order").reset_index(drop=True)


def build_monthly_pnl(fact: pd.DataFrame) -> pd.DataFrame:
    actual_fact = select_actual_fact(fact)
    if actual_fact.empty:
        return pd.DataFrame()
    pivot = actual_fact.pivot_table(
        index="period",
        columns="financial_line",
        values="amount",
        aggfunc="sum",
        fill_value=0,
    ).reset_index()
    for column in P_AND_L_ORDER:
        if column not in pivot.columns and column not in {"Gross Profit", "Operating Profit", "Net Profit"}:
            pivot[column] = 0.0
    pivot["Revenue"] = pivot.get("Revenue", 0.0) + pivot.get("Unclassified Income", 0.0)
    pivot["Gross Profit"] = pivot["Revenue"] + pivot.get("COGS", 0.0)
    pivot["Operating Profit"] = pivot["Gross Profit"] + pivot.get("Operating Expense", 0.0)
    pivot["Net Profit"] = (
        pivot["Operating Profit"]
        + pivot.get("Other Income", 0.0)
        + pivot.get("Other Expense", 0.0)
        + pivot.get("Tax", 0.0)
        + pivot.get("Unclassified", 0.0)
        + pivot.get("Unclassified Expense", 0.0)
    )
    pivot["Gross Margin"] = pivot["Gross Profit"] / pivot["Revenue"].replace(0, np.nan)
    pivot["Net Margin"] = pivot["Net Profit"] / pivot["Revenue"].replace(0, np.nan)
    keep_columns = [
        "period",
        "Revenue",
        "COGS",
        "Gross Profit",
        "Operating Expense",
        "Operating Profit",
        "Other Income",
        "Other Expense",
        "Tax",
        "Net Profit",
        "Gross Margin",
        "Net Margin",
    ]
    keep_columns = [column for column in keep_columns if column in pivot.columns]
    return pivot[keep_columns].sort_values("period").reset_index(drop=True)


def build_variance_summary(fact: pd.DataFrame) -> pd.DataFrame:
    if "scenario" not in fact.columns:
        return pd.DataFrame()
    scenarios = set(fact["scenario"].astype("string").str.lower().dropna())
    if "actual" not in scenarios:
        return pd.DataFrame()
    comparison = None
    for candidate in ("budget", "forecast", "prior"):
        if candidate in scenarios:
            comparison = candidate
            break
    if comparison is None:
        return pd.DataFrame()

    summary = fact.copy()
    summary["scenario_key"] = summary["scenario"].astype("string").str.lower()
    pivot = summary.pivot_table(
        index=["financial_line"],
        columns="scenario_key",
        values="amount",
        aggfunc="sum",
        fill_value=0,
    ).reset_index()
    pivot["comparison"] = comparison.title()
    pivot["variance"] = pivot["actual"] - pivot[comparison]
    pivot["variance_percent"] = pivot["variance"] / pivot[comparison].abs().replace(0, np.nan)
    pivot = pivot.sort_values("variance", key=lambda series: series.abs(), ascending=False).reset_index(drop=True)
    return pivot


def summarize_by_dimension(fact: pd.DataFrame, dimension: str, top_n: int) -> pd.DataFrame:
    actual_fact = select_actual_fact(fact)
    if dimension not in actual_fact.columns or actual_fact.empty:
        return pd.DataFrame()
    grouped = actual_fact.groupby([dimension, "financial_line"], dropna=False)["amount"].sum().unstack(fill_value=0)
    grouped["Revenue"] = grouped.get("Revenue", 0.0) + grouped.get("Unclassified Income", 0.0)
    grouped["Gross Profit"] = grouped["Revenue"] + grouped.get("COGS", 0.0)
    grouped["Operating Profit"] = grouped["Gross Profit"] + grouped.get("Operating Expense", 0.0)
    grouped["Net Profit"] = (
        grouped["Operating Profit"]
        + grouped.get("Other Income", 0.0)
        + grouped.get("Other Expense", 0.0)
        + grouped.get("Tax", 0.0)
        + grouped.get("Unclassified", 0.0)
        + grouped.get("Unclassified Expense", 0.0)
    )
    grouped["Gross Margin"] = grouped["Gross Profit"] / grouped["Revenue"].replace(0, np.nan)
    grouped["Net Margin"] = grouped["Net Profit"] / grouped["Revenue"].replace(0, np.nan)
    return grouped.sort_values("Revenue", ascending=False).head(top_n).reset_index()


def format_number(value: object, decimals: int = 0) -> str:
    if value is None or pd.isna(value):
        return "n/a"
    return f"{float(value):,.{decimals}f}"


def format_currency(value: object, currency_symbol: str) -> str:
    if value is None or pd.isna(value):
        return "n/a"
    value = float(value)
    sign = "-" if value < 0 else ""
    return f"{sign}{currency_symbol}{abs(value):,.2f}"


def format_percent(value: object) -> str:
    if value is None or pd.isna(value):
        return "n/a"
    return f"{float(value) * 100:.1f}%"


def markdown_table(headers: list[str], rows: list[list[object]]) -> str:
    if not rows:
        return "_No suitable data available._"
    header_line = "| " + " | ".join(headers) + " |"
    separator_line = "| " + " | ".join("---" for _ in headers) + " |"
    body_lines = ["| " + " | ".join(str(cell) for cell in row) + " |" for row in rows]
    return "\n".join([header_line, separator_line, *body_lines])


def safe_filename(text: str) -> str:
    return re.sub(r"[^0-9a-zA-Z_-]+", "_", text).strip("_").lower()


def wrap_labels(labels: Iterable[object], width: int = 24) -> list[str]:
    return [textwrap.fill(str(label), width=width) for label in labels]


def apply_chart_style() -> None:
    try:
        plt.style.use("seaborn-v0_8-whitegrid")
    except OSError:
        plt.style.use("default")


def save_figure(fig: plt.Figure, path: Path) -> Path:
    fig.tight_layout()
    fig.savefig(path, dpi=150, bbox_inches="tight")
    plt.close(fig)
    return path


def create_charts(
    fact: pd.DataFrame,
    monthly_pnl: pd.DataFrame,
    variance: pd.DataFrame,
    roles: FinancialRoles,
    charts_dir: Path,
    top_n: int,
) -> tuple[list[Path], list[str]]:
    charts_dir.mkdir(parents=True, exist_ok=True)
    apply_chart_style()
    chart_paths: list[Path] = []
    chart_errors: list[str] = []

    def attempt(name: str, func) -> None:
        try:
            path = func()
            if path:
                chart_paths.append(path)
        except Exception as exc:
            chart_errors.append(f"{name}: {exc}")

    attempt("monthly_pnl_trend", lambda: plot_monthly_pnl_trend(monthly_pnl, charts_dir))
    attempt("pnl_waterfall", lambda: plot_pnl_waterfall(fact, charts_dir))
    attempt("expense_drivers", lambda: plot_expense_drivers(fact, charts_dir, top_n))
    attempt("margin_trend", lambda: plot_margin_trend(monthly_pnl, charts_dir))
    if not variance.empty:
        attempt("variance", lambda: plot_variance(variance, charts_dir, top_n))
    if roles.dimensions:
        attempt(
            f"{roles.dimensions[0]}_profitability",
            lambda: plot_dimension_profitability(fact, roles.dimensions[0], charts_dir, top_n),
        )
    return chart_paths, chart_errors


def plot_monthly_pnl_trend(monthly_pnl: pd.DataFrame, charts_dir: Path) -> Path | None:
    if monthly_pnl.empty or "Revenue" not in monthly_pnl:
        return None
    monthly_pnl = monthly_pnl[monthly_pnl["period"].astype(str) != "Unassigned Period"].copy()
    if monthly_pnl.empty:
        return None
    fig, ax = plt.subplots(figsize=(11, 5))
    x = monthly_pnl["period"].astype(str)
    ax.plot(x, monthly_pnl["Revenue"], marker="o", linewidth=2, label="Revenue")
    if "Gross Profit" in monthly_pnl:
        ax.plot(x, monthly_pnl["Gross Profit"], marker="o", linewidth=2, label="Gross Profit")
    if "Net Profit" in monthly_pnl:
        ax.plot(x, monthly_pnl["Net Profit"], marker="o", linewidth=2, label="Net Profit")
    ax.set_title("Monthly P&L Trend")
    ax.set_xlabel("Period")
    ax.set_ylabel("Amount")
    ax.legend()
    ax.tick_params(axis="x", rotation=35)
    return save_figure(fig, charts_dir / "monthly_pnl_trend.png")


def plot_margin_trend(monthly_pnl: pd.DataFrame, charts_dir: Path) -> Path | None:
    if monthly_pnl.empty or "Gross Margin" not in monthly_pnl:
        return None
    monthly_pnl = monthly_pnl[monthly_pnl["period"].astype(str) != "Unassigned Period"].copy()
    if monthly_pnl.empty:
        return None
    fig, ax = plt.subplots(figsize=(10, 4.8))
    x = monthly_pnl["period"].astype(str)
    ax.plot(x, monthly_pnl["Gross Margin"] * 100, marker="o", linewidth=2, label="Gross Margin")
    if "Net Margin" in monthly_pnl:
        ax.plot(x, monthly_pnl["Net Margin"] * 100, marker="o", linewidth=2, label="Net Margin")
    ax.set_title("Margin Trend")
    ax.set_xlabel("Period")
    ax.set_ylabel("Margin (%)")
    ax.legend()
    ax.tick_params(axis="x", rotation=35)
    return save_figure(fig, charts_dir / "margin_trend.png")


def plot_pnl_waterfall(fact: pd.DataFrame, charts_dir: Path) -> Path | None:
    metrics = pnl_metrics(fact)
    rows = [
        ("Revenue", metrics["Revenue"]),
        ("COGS", metrics["COGS"]),
        ("Gross Profit", metrics["Gross Profit"]),
        ("Operating Expense", metrics["Operating Expense"]),
        ("Operating Profit", metrics["Operating Profit"]),
        ("Net Profit", metrics["Net Profit"]),
    ]
    rows = [(label, value) for label, value in rows if abs(value) > 0.000001 or label in {"Revenue", "Net Profit"}]
    if not rows:
        return None

    labels = [row[0] for row in rows]
    values = [row[1] for row in rows]
    colors = ["#2f6f73" if value >= 0 else "#b45f4d" for value in values]
    fig, ax = plt.subplots(figsize=(10, 5))
    ax.bar(labels, values, color=colors)
    ax.axhline(0, color="#333333", linewidth=0.8)
    ax.set_title("P&L Summary")
    ax.set_ylabel("Amount")
    ax.tick_params(axis="x", rotation=25)
    for index, value in enumerate(values):
        va = "bottom" if value >= 0 else "top"
        ax.text(index, value, f"{value:,.0f}", ha="center", va=va, fontsize=9)
    return save_figure(fig, charts_dir / "pnl_summary.png")


def plot_expense_drivers(fact: pd.DataFrame, charts_dir: Path, top_n: int) -> Path | None:
    actual_fact = select_actual_fact(fact)
    expense_fact = actual_fact[actual_fact["amount"] < 0].copy()
    if expense_fact.empty:
        return None
    group_column = "account" if "account" in expense_fact and expense_fact["account"].nunique() > 1 else "financial_line"
    grouped = expense_fact.groupby(group_column)["amount"].sum().abs().sort_values(ascending=False).head(top_n).sort_values()
    if grouped.empty:
        return None
    fig, ax = plt.subplots(figsize=(10, max(4.5, len(grouped) * 0.55)))
    y_positions = np.arange(len(grouped))
    ax.barh(y_positions, grouped.values, color="#7c5c2e")
    ax.set_yticks(y_positions)
    ax.set_yticklabels(wrap_labels(grouped.index))
    ax.set_title(f"Top {top_n} Expense Drivers")
    ax.set_xlabel("Expense Amount")
    for index, value in enumerate(grouped.values):
        ax.text(value, index, f" {value:,.0f}", va="center", fontsize=9)
    return save_figure(fig, charts_dir / "top_expense_drivers.png")


def plot_variance(variance: pd.DataFrame, charts_dir: Path, top_n: int) -> Path | None:
    if variance.empty or "variance" not in variance:
        return None
    grouped = variance.set_index("financial_line")["variance"].sort_values(key=lambda series: series.abs(), ascending=False)
    grouped = grouped.head(top_n).sort_values()
    if grouped.empty:
        return None
    fig, ax = plt.subplots(figsize=(10, max(4.5, len(grouped) * 0.55)))
    y_positions = np.arange(len(grouped))
    colors = ["#2f6f73" if value >= 0 else "#b45f4d" for value in grouped.values]
    ax.barh(y_positions, grouped.values, color=colors)
    ax.axvline(0, color="#333333", linewidth=0.8)
    ax.set_yticks(y_positions)
    ax.set_yticklabels(wrap_labels(grouped.index))
    ax.set_title("Actual vs Comparison Variance")
    ax.set_xlabel("Variance")
    return save_figure(fig, charts_dir / "actual_vs_comparison_variance.png")


def plot_dimension_profitability(fact: pd.DataFrame, dimension: str, charts_dir: Path, top_n: int) -> Path | None:
    summary = summarize_by_dimension(fact, dimension, top_n)
    if summary.empty or "Revenue" not in summary:
        return None
    summary = summary.sort_values("Revenue")
    fig, ax = plt.subplots(figsize=(10, max(4.5, len(summary) * 0.55)))
    y_positions = np.arange(len(summary))
    ax.barh(y_positions, summary["Revenue"], color="#4d72b0", label="Revenue")
    if "Net Profit" in summary:
        ax.barh(y_positions, summary["Net Profit"], color="#2f6f73", label="Net Profit")
    ax.set_yticks(y_positions)
    ax.set_yticklabels(wrap_labels(summary[dimension]))
    ax.set_title(f"Profitability by {dimension.replace('_', ' ').title()}")
    ax.set_xlabel("Amount")
    ax.legend()
    return save_figure(fig, charts_dir / f"profitability_by_{safe_filename(dimension)}.png")


def build_kpi_rows(fact: pd.DataFrame, currency_symbol: str) -> list[list[str]]:
    metrics = pnl_metrics(fact)
    actual_fact = select_actual_fact(fact)
    rows = [
        ["Finance fact rows", format_number(len(fact)), "Normalized transaction/line-item rows"],
        ["Revenue", format_currency(metrics["Revenue"], currency_symbol), "Top-line sales or income"],
        ["Gross profit", format_currency(metrics["Gross Profit"], currency_symbol), "Revenue after COGS"],
        ["Gross margin", format_percent(metrics["Gross Margin"]), "Profitability before operating expenses"],
        ["Operating profit", format_currency(metrics["Operating Profit"], currency_symbol), "Core business profit"],
        ["Net profit", format_currency(metrics["Net Profit"], currency_symbol), "Bottom-line profit after all classified lines"],
        ["Net margin", format_percent(metrics["Net Margin"]), "Bottom-line profitability rate"],
        ["Total expenses", format_currency(metrics["Total Expenses"], currency_symbol), "COGS plus operating/other expenses"],
    ]
    if "period" in actual_fact and actual_fact["period"].nunique() > 0:
        valid_periods = actual_fact[actual_fact["period"].astype(str) != "Unassigned Period"]["period"].nunique()
        rows.append(["Periods", format_number(valid_periods), "Number of dated reporting periods"])
    return rows


def build_insights(
    fact: pd.DataFrame,
    monthly_pnl: pd.DataFrame,
    variance: pd.DataFrame,
    roles: FinancialRoles,
    currency_symbol: str,
    top_n: int,
) -> list[str]:
    metrics = pnl_metrics(fact)
    insights = [
        f"Revenue is {format_currency(metrics['Revenue'], currency_symbol)} with gross margin of {format_percent(metrics['Gross Margin'])} and net margin of {format_percent(metrics['Net Margin'])}.",
        f"Net profit is {format_currency(metrics['Net Profit'], currency_symbol)} after total expenses of {format_currency(metrics['Total Expenses'], currency_symbol)}.",
    ]

    dated_monthly = monthly_pnl[monthly_pnl["period"].astype(str) != "Unassigned Period"].copy()
    if not dated_monthly.empty and len(dated_monthly) >= 2:
        first = dated_monthly.iloc[0]
        last = dated_monthly.iloc[-1]
        if first["Revenue"] != 0:
            revenue_change = (last["Revenue"] - first["Revenue"]) / abs(first["Revenue"])
            direction = "increased" if revenue_change >= 0 else "decreased"
            insights.append(
                f"Revenue {direction} by {abs(revenue_change) * 100:.1f}% from {first['period']} to {last['period']} "
                f"({format_currency(first['Revenue'], currency_symbol)} -> {format_currency(last['Revenue'], currency_symbol)})."
            )
        if pd.notna(first.get("Gross Margin")) and pd.notna(last.get("Gross Margin")):
            margin_delta = (last["Gross Margin"] - first["Gross Margin"]) * 100
            direction = "expanded" if margin_delta >= 0 else "compressed"
            insights.append(f"Gross margin {direction} by {abs(margin_delta):.1f} percentage points over the reporting window.")

    actual_fact = select_actual_fact(fact)
    expense_fact = actual_fact[actual_fact["amount"] < 0]
    if not expense_fact.empty:
        expense_by_line = expense_fact.groupby("financial_line")["amount"].sum().abs().sort_values(ascending=False)
        top_expense_line = expense_by_line.index[0]
        top_expense_value = expense_by_line.iloc[0]
        insights.append(
            f"The largest expense bucket is {top_expense_line} at {format_currency(top_expense_value, currency_symbol)}, "
            "which should be reviewed for controllability and trend risk."
        )

    for dimension in roles.dimensions[:2]:
        summary = summarize_by_dimension(fact, dimension, top_n)
        if summary.empty or "Revenue" not in summary:
            continue
        top_row = summary.iloc[0]
        revenue_share = top_row["Revenue"] / metrics["Revenue"] if metrics["Revenue"] else np.nan
        insights.append(
            f"The top {dimension.replace('_', ' ')} is '{top_row[dimension]}' with "
            f"{format_currency(top_row['Revenue'], currency_symbol)}, representing {format_percent(revenue_share)} of revenue."
        )
        if "Net Margin" in summary and summary["Net Margin"].notna().any():
            low_margin = summary.sort_values("Net Margin").iloc[0]
            insights.append(
                f"'{low_margin[dimension]}' has the weakest net margin among top {dimension.replace('_', ' ')} groups "
                f"({format_percent(low_margin['Net Margin'])}); review pricing, cost allocation, and mix."
            )

    if not variance.empty:
        top_variance = variance.iloc[0]
        comparison = top_variance.get("comparison", "Comparison")
        insights.append(
            f"Largest Actual vs {comparison} variance is {top_variance['financial_line']} at "
            f"{format_currency(top_variance['variance'], currency_symbol)} ({format_percent(top_variance['variance_percent'])})."
        )

    return insights[:10]


def build_recommendations(fact: pd.DataFrame, roles: FinancialRoles) -> list[str]:
    recommendations = [
        "Validate the financial sign convention with Finance: this template reports income as positive and costs/expenses as negative.",
        "Lock a data dictionary for account, account type, scenario, and cost center fields to reduce manual mapping risk.",
    ]
    metrics = pnl_metrics(fact)
    if pd.notna(metrics["Gross Margin"]) and metrics["Gross Margin"] < 0.35:
        recommendations.append("Prioritize gross margin review: analyze COGS, supplier pricing, shrinkage, and product/service mix.")
    else:
        recommendations.append("Use gross margin by product, store, department, or region to find where profitable growth can scale.")
    if roles.dimensions:
        recommendations.append(
            f"Build a recurring driver view by {roles.dimensions[0].replace('_', ' ')} to separate revenue growth from margin dilution."
        )
    if "scenario" in fact.columns and fact["scenario"].astype("string").str.lower().isin(["budget", "forecast"]).any():
        recommendations.append("Turn variance output into a monthly close pack with owner, root cause, and action status columns.")
    else:
        recommendations.append("Add Budget or Forecast data to enable variance analysis and a more complete finance review cadence.")
    return recommendations[:6]


def build_data_quality_rows(profile: dict[str, object], source_info: dict[str, object], fact: pd.DataFrame) -> list[list[str]]:
    original_shape = profile.get("original_shape", (0, 0))
    cleaned_shape = profile.get("cleaned_shape", (0, 0))
    rows = [
        ["Original shape", f"{original_shape[0]:,} rows x {original_shape[1]:,} columns"],
        ["Cleaned shape", f"{cleaned_shape[0]:,} rows x {cleaned_shape[1]:,} columns"],
        ["Finance fact rows", format_number(len(fact))],
        ["Amount source", str(source_info.get("amount_source", "not detected"))],
        ["Duplicate rows removed", format_number(profile.get("duplicate_rows_removed", 0))],
    ]
    parsed_dates = profile.get("parsed_date_columns", {})
    if isinstance(parsed_dates, dict) and parsed_dates:
        rows.append(["Date columns parsed", ", ".join(f"{column} ({ratio:.0%})" for column, ratio in parsed_dates.items())])
    converted_numeric = profile.get("converted_numeric_columns", {})
    if isinstance(converted_numeric, dict) and converted_numeric:
        rows.append(
            ["Numeric columns converted", ", ".join(f"{column} ({ratio:.0%})" for column, ratio in converted_numeric.items())]
        )
    filled_missing = profile.get("filled_missing_values", {})
    if isinstance(filled_missing, dict) and filled_missing:
        rows.append(["Missing values handled", "; ".join(f"{column}: {rule}" for column, rule in filled_missing.items())])
    unclassified_count = int(fact["financial_line"].isin(["Unclassified", "Unclassified Income", "Unclassified Expense"]).sum())
    rows.append(["Unclassified finance rows", format_number(unclassified_count)])
    return rows


def dataframe_to_markdown_rows(df: pd.DataFrame, currency_symbol: str, max_rows: int = 12) -> tuple[list[str], list[list[str]]]:
    if df.empty:
        return [], []
    display = df.head(max_rows).copy()
    headers = [str(column).replace("_", " ").title() for column in display.columns]
    rows: list[list[str]] = []
    for _, row in display.iterrows():
        formatted_row: list[str] = []
        for column, value in row.items():
            column_text = str(column).lower()
            if "margin" in column_text or "percent" in column_text:
                formatted_row.append(format_percent(value))
            elif isinstance(value, (int, float, np.integer, np.floating)) and not pd.isna(value):
                formatted_row.append(format_currency(value, currency_symbol))
            else:
                formatted_row.append("n/a" if pd.isna(value) else str(value))
        rows.append(formatted_row)
    return headers, rows


def build_report(
    *,
    input_path: Path,
    cleaned_path: Path,
    fact_path: Path,
    workbook_path: Path,
    fact: pd.DataFrame,
    pnl_summary: pd.DataFrame,
    monthly_pnl: pd.DataFrame,
    variance: pd.DataFrame,
    roles: FinancialRoles,
    profile: dict[str, object],
    source_info: dict[str, object],
    chart_paths: list[Path],
    chart_errors: list[str],
    args: argparse.Namespace,
) -> str:
    generated_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    report_lines: list[str] = [
        f"# Financial Report - {input_path.name}",
        "",
        f"Generated at: {generated_at}",
        "",
        "## Executive Summary",
        "",
    ]
    report_lines.extend(
        f"- {insight}"
        for insight in build_insights(fact, monthly_pnl, variance, roles, args.currency_symbol, args.top_n)
    )

    report_lines.extend(
        [
            "",
            "## KPI Snapshot",
            "",
            markdown_table(["KPI", "Value", "Why it matters"], build_kpi_rows(fact, args.currency_symbol)),
            "",
            "## P&L Summary",
            "",
        ]
    )
    pnl_headers, pnl_rows = dataframe_to_markdown_rows(pnl_summary, args.currency_symbol)
    report_lines.append(markdown_table(pnl_headers, pnl_rows))

    report_lines.extend(["", "## Monthly Performance", ""])
    monthly_headers, monthly_rows = dataframe_to_markdown_rows(monthly_pnl, args.currency_symbol, max_rows=18)
    report_lines.append(markdown_table(monthly_headers, monthly_rows))

    if not variance.empty:
        report_lines.extend(["", "## Variance Analysis", ""])
        variance_headers, variance_rows = dataframe_to_markdown_rows(variance, args.currency_symbol, max_rows=args.top_n)
        report_lines.append(markdown_table(variance_headers, variance_rows))

    if roles.dimensions:
        report_lines.extend(["", "## Driver Analysis", ""])
        for dimension in roles.dimensions[:3]:
            summary = summarize_by_dimension(fact, dimension, args.top_n)
            if summary.empty:
                continue
            report_lines.extend([f"### By {dimension.replace('_', ' ').title()}", ""])
            headers, rows = dataframe_to_markdown_rows(summary, args.currency_symbol, max_rows=args.top_n)
            report_lines.extend([markdown_table(headers, rows), ""])

    report_lines.extend(
        [
            "",
            "## Data Quality And Controls",
            "",
            markdown_table(["Check", "Result"], build_data_quality_rows(profile, source_info, fact)),
            "",
            "## Recommendations",
            "",
        ]
    )
    report_lines.extend(f"- {recommendation}" for recommendation in build_recommendations(fact, roles))

    report_lines.extend(
        [
            "",
            "## Output Files",
            "",
            f"- Cleaned CSV: `{cleaned_path.as_posix()}`",
            f"- Finance fact CSV: `{fact_path.as_posix()}`",
            f"- Excel workbook: `{workbook_path.as_posix()}`",
        ]
    )
    if chart_paths:
        report_lines.append("- Charts:")
        report_lines.extend(f"  - `{path.as_posix()}`" for path in chart_paths)
    else:
        report_lines.append("- Charts: no chart type matched the current schema.")

    if chart_errors:
        report_lines.extend(["", "## Chart Warnings", ""])
        report_lines.extend(f"- {error}" for error in chart_errors)

    report_lines.extend(
        [
            "",
            "## Assumptions",
            "",
            "- P&L sign convention is income positive and costs/expenses negative unless `--sign-mode as-is` is used.",
            "- If the CSV has separate revenue and cost columns, the template creates two finance fact rows per source row.",
            "- If Actual/Budget/Forecast/Prior columns exist, they are melted into scenario rows for variance analysis.",
            "- Auto-inference is a starting point; use CLI column overrides for production reporting.",
            "",
        ]
    )
    return "\n".join(report_lines)


def build_workbook(
    workbook_path: Path,
    cleaned_df: pd.DataFrame,
    fact: pd.DataFrame,
    pnl_summary: pd.DataFrame,
    monthly_pnl: pd.DataFrame,
    variance: pd.DataFrame,
    quality_rows: list[list[str]],
) -> None:
    workbook_path.parent.mkdir(parents=True, exist_ok=True)
    quality_df = pd.DataFrame(quality_rows, columns=["check", "result"])
    with pd.ExcelWriter(workbook_path, engine="openpyxl") as writer:
        pnl_summary.to_excel(writer, sheet_name="P&L Summary", index=False)
        monthly_pnl.to_excel(writer, sheet_name="Monthly P&L", index=False)
        if not variance.empty:
            variance.to_excel(writer, sheet_name="Variance", index=False)
        fact.to_excel(writer, sheet_name="Finance Fact", index=False)
        cleaned_df.to_excel(writer, sheet_name="Cleaned Data", index=False)
        quality_df.to_excel(writer, sheet_name="Data Quality", index=False)
    format_workbook(workbook_path)


def format_workbook(workbook_path: Path) -> None:
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        return

    workbook = load_workbook(workbook_path)
    header_fill = PatternFill("solid", fgColor="1F4E5F")
    header_font = Font(color="FFFFFF", bold=True)
    for worksheet in workbook.worksheets:
        worksheet.freeze_panes = "A2"
        if worksheet.max_row and worksheet.max_column:
            worksheet.auto_filter.ref = worksheet.dimensions
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
        for column_cells in worksheet.columns:
            values = [str(cell.value) if cell.value is not None else "" for cell in column_cells]
            max_length = max(len(value) for value in values) if values else 10
            width = min(max(max_length + 2, 12), 45)
            worksheet.column_dimensions[column_cells[0].column_letter].width = width
    workbook.save(workbook_path)


def save_outputs(
    *,
    input_path: Path,
    output_dir: Path,
    cleaned_df: pd.DataFrame,
    fact: pd.DataFrame,
    report: str,
    pnl_summary: pd.DataFrame,
    monthly_pnl: pd.DataFrame,
    variance: pd.DataFrame,
    quality_rows: list[list[str]],
) -> tuple[Path, Path, Path, Path]:
    output_dir.mkdir(parents=True, exist_ok=True)
    stem = safe_filename(input_path.stem)
    cleaned_path = output_dir / f"{stem}_cleaned.csv"
    fact_path = output_dir / f"{stem}_finance_fact.csv"
    report_path = output_dir / f"{stem}_financial_report.md"
    workbook_path = output_dir / f"{stem}_financial_workbook.xlsx"
    cleaned_df.to_csv(cleaned_path, index=False)
    fact.to_csv(fact_path, index=False)
    report_path.write_text(report, encoding="utf-8")
    build_workbook(workbook_path, cleaned_df, fact, pnl_summary, monthly_pnl, variance, quality_rows)
    return cleaned_path, fact_path, report_path, workbook_path


def resolve_input_path(raw_path: str) -> Path:
    path = Path(raw_path)
    if path.exists():
        return path
    script_relative = Path(__file__).resolve().parent / raw_path
    if script_relative.exists():
        return script_relative
    raise FileNotFoundError(f"Input file not found: {raw_path}")


def validate_args(args: argparse.Namespace) -> None:
    if not 1 <= args.fiscal_year_start <= 12:
        raise ValueError("--fiscal-year-start must be between 1 and 12.")


def main() -> None:
    args = parse_args()
    validate_args(args)
    input_path = resolve_input_path(args.input)
    output_dir = Path(args.output_dir)
    if not output_dir.is_absolute():
        output_dir = Path.cwd() / output_dir

    raw_df = load_csv(input_path, args.encoding)
    cleaned_df, profile = clean_dataframe(raw_df, title_case_text=not args.no_title_case)
    roles = infer_roles(cleaned_df, args)
    fact, source_info = build_finance_fact(cleaned_df, roles, args)
    pnl_summary = build_pnl_summary(fact)
    monthly_pnl = build_monthly_pnl(fact)
    variance = build_variance_summary(fact)

    charts_dir = output_dir / "charts"
    chart_paths, chart_errors = create_charts(fact, monthly_pnl, variance, roles, charts_dir, args.top_n)

    stem = safe_filename(input_path.stem)
    cleaned_path = output_dir / f"{stem}_cleaned.csv"
    fact_path = output_dir / f"{stem}_finance_fact.csv"
    workbook_path = output_dir / f"{stem}_financial_workbook.xlsx"
    quality_rows = build_data_quality_rows(profile, source_info, fact)
    report = build_report(
        input_path=input_path,
        cleaned_path=cleaned_path,
        fact_path=fact_path,
        workbook_path=workbook_path,
        fact=fact,
        pnl_summary=pnl_summary,
        monthly_pnl=monthly_pnl,
        variance=variance,
        roles=roles,
        profile=profile,
        source_info=source_info,
        chart_paths=chart_paths,
        chart_errors=chart_errors,
        args=args,
    )
    cleaned_path, fact_path, report_path, workbook_path = save_outputs(
        input_path=input_path,
        output_dir=output_dir,
        cleaned_df=cleaned_df,
        fact=fact,
        report=report,
        pnl_summary=pnl_summary,
        monthly_pnl=monthly_pnl,
        variance=variance,
        quality_rows=quality_rows,
    )

    print("Done.")
    print(f"Cleaned CSV: {cleaned_path}")
    print(f"Finance fact CSV: {fact_path}")
    print(f"Report: {report_path}")
    print(f"Excel workbook: {workbook_path}")
    print(f"Charts: {charts_dir}")
    if chart_errors:
        print("Chart warnings:")
        for error in chart_errors:
            print(f"- {error}")


if __name__ == "__main__":
    main()
