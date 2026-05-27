"""
Create a retail sales report from retail_sales.csv.

Outputs:
    retail_report_output/retail_sales_cleaned.csv
    retail_report_output/retail_sales_report.md
    retail_report_output/retail_sales_report.html
    retail_report_output/retail_sales_report.xlsx
    retail_report_output/charts/*.png
"""

from __future__ import annotations

import argparse
import html
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

import matplotlib

matplotlib.use("Agg")

import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows


REQUIRED_COLUMNS = ["OrderID", "Date", "Region", "Category", "CustomerSegment", "Sales", "Profit"]


@dataclass(frozen=True)
class ReportOutputs:
    cleaned_csv: Path
    markdown_report: Path
    html_report: Path
    workbook: Path
    charts: list[Path]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Generate a retail sales dashboard, charts, and insights from retail_sales.csv.",
        formatter_class=argparse.ArgumentDefaultsHelpFormatter,
    )
    parser.add_argument("--input", default="retail_sales.csv", help="Input CSV path.")
    parser.add_argument("--output-dir", default="retail_report_output", help="Output directory.")
    parser.add_argument("--currency-symbol", default="$", help="Currency symbol used in reports.")
    return parser.parse_args()


def money(value: float, symbol: str = "$") -> str:
    return f"{symbol}{value:,.0f}"


def number(value: float) -> str:
    return f"{value:,.0f}"


def percent(value: float) -> str:
    return f"{value:.1%}"


def load_and_clean(input_path: Path) -> tuple[pd.DataFrame, pd.DataFrame]:
    raw = pd.read_csv(input_path)
    raw.columns = [str(column).strip() for column in raw.columns]

    missing_columns = [column for column in REQUIRED_COLUMNS if column not in raw.columns]
    if missing_columns:
        raise ValueError(f"Missing required columns: {', '.join(missing_columns)}")

    quality_rows: list[dict[str, object]] = []
    quality_rows.append({"Metric": "Rows before cleaning", "Value": len(raw), "Notes": "Original CSV rows"})
    quality_rows.append({"Metric": "Columns", "Value": len(raw.columns), "Notes": ", ".join(raw.columns)})

    df = raw.copy()
    duplicate_count = int(df.duplicated().sum())
    df = df.drop_duplicates().copy()
    quality_rows.append({"Metric": "Duplicate rows removed", "Value": duplicate_count, "Notes": "Exact duplicate rows"})

    for column in ["Region", "Category", "CustomerSegment"]:
        df[column] = df[column].astype("string").str.strip().str.replace(r"\s+", " ", regex=True)

    df["Date"] = pd.to_datetime(df["Date"], errors="coerce")
    df["Sales"] = pd.to_numeric(df["Sales"], errors="coerce")
    df["Profit"] = pd.to_numeric(df["Profit"], errors="coerce")
    df["OrderID"] = pd.to_numeric(df["OrderID"], errors="coerce").astype("Int64")

    for column in REQUIRED_COLUMNS:
        quality_rows.append(
            {
                "Metric": f"Missing values - {column}",
                "Value": int(df[column].isna().sum()),
                "Notes": "After type conversion",
            }
        )

    df = df.dropna(subset=["OrderID", "Date", "Region", "Category", "CustomerSegment", "Sales", "Profit"]).copy()
    df = df.sort_values(["Date", "OrderID"]).reset_index(drop=True)
    df["ProfitMargin"] = np.where(df["Sales"] != 0, df["Profit"] / df["Sales"], np.nan)
    df["Month"] = df["Date"].dt.strftime("%Y-%m")
    df["DayName"] = df["Date"].dt.day_name()

    quality_rows.extend(
        [
            {"Metric": "Rows after cleaning", "Value": len(df), "Notes": "Rows with required fields"},
            {
                "Metric": "Date range",
                "Value": f"{df['Date'].min().date()} to {df['Date'].max().date()}",
                "Notes": "Parsed from Date column",
            },
            {
                "Metric": "Overall profit margin",
                "Value": percent(df["Profit"].sum() / df["Sales"].sum()) if df["Sales"].sum() else "n/a",
                "Notes": "Total Profit / Total Sales",
            },
        ]
    )

    return df, pd.DataFrame(quality_rows)


def summarize_group(df: pd.DataFrame, column: str) -> pd.DataFrame:
    summary = (
        df.groupby(column, dropna=False)
        .agg(Orders=("OrderID", "count"), Sales=("Sales", "sum"), Profit=("Profit", "sum"))
        .reset_index()
        .sort_values("Sales", ascending=False)
    )
    summary["ProfitMargin"] = summary["Profit"] / summary["Sales"]
    summary["SalesShare"] = summary["Sales"] / summary["Sales"].sum()
    summary["ProfitShare"] = summary["Profit"] / summary["Profit"].sum()
    return summary


def build_summaries(df: pd.DataFrame) -> dict[str, pd.DataFrame]:
    daily = (
        df.groupby("Date")
        .agg(Orders=("OrderID", "count"), Sales=("Sales", "sum"), Profit=("Profit", "sum"))
        .reset_index()
        .sort_values("Date")
    )
    daily["ProfitMargin"] = daily["Profit"] / daily["Sales"]

    return {
        "daily": daily,
        "region": summarize_group(df, "Region"),
        "category": summarize_group(df, "Category"),
        "segment": summarize_group(df, "CustomerSegment"),
    }


def build_kpis(df: pd.DataFrame, summaries: dict[str, pd.DataFrame]) -> dict[str, object]:
    total_sales = float(df["Sales"].sum())
    total_profit = float(df["Profit"].sum())
    return {
        "orders": int(df["OrderID"].nunique()),
        "total_sales": total_sales,
        "total_profit": total_profit,
        "profit_margin": total_profit / total_sales if total_sales else 0,
        "avg_order_value": total_sales / len(df) if len(df) else 0,
        "best_region": str(summaries["region"].iloc[0]["Region"]),
        "best_category": str(summaries["category"].iloc[0]["Category"]),
        "best_segment": str(summaries["segment"].iloc[0]["CustomerSegment"]),
        "date_start": df["Date"].min(),
        "date_end": df["Date"].max(),
    }


def build_insights(df: pd.DataFrame, summaries: dict[str, pd.DataFrame], kpis: dict[str, object], symbol: str) -> list[str]:
    category = summaries["category"].iloc[0]
    region = summaries["region"].iloc[0]
    segment = summaries["segment"].iloc[0]
    margin_segment = summaries["segment"].sort_values("ProfitMargin", ascending=False).iloc[0]
    weakest_category = summaries["category"].sort_values("ProfitMargin").iloc[0]

    first_five = df.head(5)
    last_five = df.tail(5)
    sales_delta = (last_five["Sales"].sum() / first_five["Sales"].sum()) - 1
    profit_delta = (last_five["Profit"].sum() / first_five["Profit"].sum()) - 1

    south = summaries["region"][summaries["region"]["Region"].eq("South")]
    south_text = ""
    if not south.empty:
        row = south.iloc[0]
        south_text = (
            f" South is the weakest region, with {money(row['Sales'], symbol)} in sales "
            f"and a {percent(row['ProfitMargin'])} profit margin."
        )

    return [
        (
            f"Total sales reached {money(kpis['total_sales'], symbol)} across {number(kpis['orders'])} orders; "
            f"profit reached {money(kpis['total_profit'], symbol)}, equal to a "
            f"{percent(kpis['profit_margin'])} profit margin."
        ),
        (
            f"The {category['Category']} category is the primary driver: {money(category['Sales'], symbol)} in sales "
            f"({percent(category['SalesShare'])} of total sales) and {money(category['Profit'], symbol)} in profit."
        ),
        (
            f"The {region['Region']} region leads with {money(region['Sales'], symbol)} in sales "
            f"and {money(region['Profit'], symbol)} in profit.{south_text}"
        ),
        (
            f"The {segment['CustomerSegment']} segment generates the most sales "
            f"({money(segment['Sales'], symbol)}); {margin_segment['CustomerSegment']} has the strongest profit margin "
            f"at {percent(margin_segment['ProfitMargin'])}."
        ),
        (
            f"Sales in the final 5 days were {percent(sales_delta)} higher than the first 5 days; profit rose "
            f"{percent(profit_delta)}, indicating stronger sales momentum late in the period."
        ),
        (
            f"{weakest_category['Category']} has the lowest profit margin "
            f"({percent(weakest_category['ProfitMargin'])}); review pricing, discounts, or product mix."
        ),
    ]


def build_recommendations(summaries: dict[str, pd.DataFrame]) -> list[str]:
    top_category = summaries["category"].iloc[0]
    weakest_category = summaries["category"].sort_values("ProfitMargin").iloc[0]
    top_region = summaries["region"].iloc[0]
    weakest_region = summaries["region"].sort_values("Sales").iloc[0]
    best_margin_segment = summaries["segment"].sort_values("ProfitMargin", ascending=False).iloc[0]

    return [
        f"Prioritize sales budget for {top_category['Category']} because it contributes the most sales and profit.",
        f"Optimize {weakest_category['Category']} through bundles, pricing adjustments, or cost reductions to bring margin closer to the overall level.",
        f"Replicate the {top_region['Region']} playbook in {weakest_region['Region']}: winning categories, promotions, and customer care scripts.",
        f"Expand the {best_margin_segment['CustomerSegment']} customer base because it is the highest-margin segment.",
        "Add Quantity, Unit Cost, Discount, and Channel to the raw data for deeper margin and campaign effectiveness analysis.",
    ]


def save_chart(path: Path) -> None:
    plt.tight_layout()
    plt.savefig(path, dpi=170, bbox_inches="tight")
    plt.close()


def create_charts(summaries: dict[str, pd.DataFrame], output_dir: Path, symbol: str) -> list[Path]:
    charts_dir = output_dir / "charts"
    charts_dir.mkdir(parents=True, exist_ok=True)
    chart_paths: list[Path] = []

    plt.style.use("seaborn-v0_8-whitegrid")
    colors = {"sales": "#2563eb", "profit": "#16a34a", "margin": "#f97316", "muted": "#94a3b8"}

    daily = summaries["daily"]
    path = charts_dir / "daily_sales_profit_trend.png"
    fig, ax = plt.subplots(figsize=(10, 5))
    ax.plot(daily["Date"], daily["Sales"], marker="o", linewidth=2.3, color=colors["sales"], label="Sales")
    ax.plot(daily["Date"], daily["Profit"], marker="o", linewidth=2.3, color=colors["profit"], label="Profit")
    ax.set_title("Daily Sales and Profit Trend", fontsize=14, weight="bold")
    ax.set_ylabel(f"Amount ({symbol})")
    ax.legend()
    ax.tick_params(axis="x", rotation=35)
    save_chart(path)
    chart_paths.append(path)

    category = summaries["category"].sort_values("Sales", ascending=True)
    path = charts_dir / "category_sales_profit.png"
    fig, ax = plt.subplots(figsize=(9, 5))
    y = np.arange(len(category))
    ax.barh(y - 0.18, category["Sales"], height=0.36, color=colors["sales"], label="Sales")
    ax.barh(y + 0.18, category["Profit"], height=0.36, color=colors["profit"], label="Profit")
    ax.set_yticks(y, category["Category"])
    ax.set_title("Sales and Profit by Category", fontsize=14, weight="bold")
    ax.set_xlabel(f"Amount ({symbol})")
    ax.legend()
    save_chart(path)
    chart_paths.append(path)

    region = summaries["region"].sort_values("Sales", ascending=True)
    path = charts_dir / "region_sales_profit.png"
    fig, ax = plt.subplots(figsize=(9, 5))
    y = np.arange(len(region))
    ax.barh(y - 0.18, region["Sales"], height=0.36, color=colors["sales"], label="Sales")
    ax.barh(y + 0.18, region["Profit"], height=0.36, color=colors["profit"], label="Profit")
    ax.set_yticks(y, region["Region"])
    ax.set_title("Sales and Profit by Region", fontsize=14, weight="bold")
    ax.set_xlabel(f"Amount ({symbol})")
    ax.legend()
    save_chart(path)
    chart_paths.append(path)

    segment = summaries["segment"].sort_values("ProfitMargin", ascending=False)
    path = charts_dir / "segment_profit_margin.png"
    fig, ax = plt.subplots(figsize=(8, 5))
    bars = ax.bar(segment["CustomerSegment"], segment["ProfitMargin"], color=colors["margin"])
    ax.set_title("Profit Margin by Customer Segment", fontsize=14, weight="bold")
    ax.set_ylabel("Profit margin")
    ax.yaxis.set_major_formatter(lambda value, _: f"{value:.0%}")
    for bar in bars:
        value = bar.get_height()
        ax.text(bar.get_x() + bar.get_width() / 2, value + 0.006, f"{value:.1%}", ha="center", va="bottom")
    save_chart(path)
    chart_paths.append(path)

    path = charts_dir / "sales_mix_by_category.png"
    fig, ax = plt.subplots(figsize=(7, 5))
    ax.pie(
        summaries["category"]["Sales"],
        labels=summaries["category"]["Category"],
        autopct="%1.1f%%",
        startangle=90,
        colors=["#2563eb", "#16a34a", "#f97316"],
    )
    ax.set_title("Sales Mix by Category", fontsize=14, weight="bold")
    ax.axis("equal")
    save_chart(path)
    chart_paths.append(path)

    return chart_paths


def df_to_markdown(df: pd.DataFrame, symbol: str) -> str:
    formatted = df.copy()
    for column in ["Sales", "Profit"]:
        if column in formatted:
            formatted[column] = formatted[column].map(lambda value: money(float(value), symbol))
    for column in ["ProfitMargin", "SalesShare", "ProfitShare"]:
        if column in formatted:
            formatted[column] = formatted[column].map(lambda value: percent(float(value)))

    headers = [str(column) for column in formatted.columns]
    rows = [[str(value) for value in row] for row in formatted.to_numpy()]
    align = ["---" if column not in {"Orders", "Sales", "Profit", "ProfitMargin", "SalesShare", "ProfitShare"} else "---:" for column in headers]
    table = [
        "| " + " | ".join(headers) + " |",
        "| " + " | ".join(align) + " |",
    ]
    table.extend("| " + " | ".join(row) + " |" for row in rows)
    return "\n".join(table)


def write_markdown(
    output_path: Path,
    input_path: Path,
    kpis: dict[str, object],
    insights: list[str],
    recommendations: list[str],
    summaries: dict[str, pd.DataFrame],
    charts: list[Path],
    symbol: str,
) -> None:
    relative_charts = [path.relative_to(output_path.parent).as_posix() for path in charts]
    lines = [
        "# Retail Sales Report",
        "",
        f"- Source file: `{input_path.name}`",
        f"- Date range: {kpis['date_start'].date()} to {kpis['date_end'].date()}",
        f"- Generated at: {datetime.now().strftime('%Y-%m-%d %H:%M')}",
        "",
        "## Executive KPIs",
        "",
        "| KPI | Value |",
        "|---|---:|",
        f"| Orders | {number(kpis['orders'])} |",
        f"| Total Sales | {money(kpis['total_sales'], symbol)} |",
        f"| Total Profit | {money(kpis['total_profit'], symbol)} |",
        f"| Profit Margin | {percent(kpis['profit_margin'])} |",
        f"| Average Order Value | {money(kpis['avg_order_value'], symbol)} |",
        f"| Best Region | {kpis['best_region']} |",
        f"| Best Category | {kpis['best_category']} |",
        f"| Best Segment | {kpis['best_segment']} |",
        "",
        "## Insights",
        "",
        *[f"- {item}" for item in insights],
        "",
        "## Recommendations",
        "",
        *[f"- {item}" for item in recommendations],
        "",
        "## Visualizations",
        "",
    ]

    for chart in relative_charts:
        lines.extend([f"![{Path(chart).stem}]({chart})", ""])

    lines.extend(
        [
            "## Performance by Category",
            "",
            df_to_markdown(summaries["category"], symbol),
            "",
            "## Performance by Region",
            "",
            df_to_markdown(summaries["region"], symbol),
            "",
            "## Performance by Segment",
            "",
            df_to_markdown(summaries["segment"], symbol),
            "",
        ]
    )
    output_path.write_text("\n".join(lines), encoding="utf-8")


def html_table(df: pd.DataFrame, symbol: str) -> str:
    formatted = df.copy()
    for column in ["Sales", "Profit"]:
        if column in formatted:
            formatted[column] = formatted[column].map(lambda value: money(float(value), symbol))
    for column in ["ProfitMargin", "SalesShare", "ProfitShare"]:
        if column in formatted:
            formatted[column] = formatted[column].map(lambda value: percent(float(value)))
    return formatted.to_html(index=False, classes="data-table", border=0, escape=True)


def write_html(
    output_path: Path,
    input_path: Path,
    kpis: dict[str, object],
    insights: list[str],
    recommendations: list[str],
    summaries: dict[str, pd.DataFrame],
    charts: list[Path],
    symbol: str,
) -> None:
    chart_html = "\n".join(
        f'<figure><img src="{html.escape(path.relative_to(output_path.parent).as_posix())}" alt="{html.escape(path.stem)}"></figure>'
        for path in charts
    )
    kpi_cards = [
        ("Orders", number(kpis["orders"])),
        ("Total Sales", money(kpis["total_sales"], symbol)),
        ("Total Profit", money(kpis["total_profit"], symbol)),
        ("Profit Margin", percent(kpis["profit_margin"])),
        ("Avg Order Value", money(kpis["avg_order_value"], symbol)),
        ("Best Category", str(kpis["best_category"])),
    ]
    kpi_html = "\n".join(f"<div><span>{html.escape(label)}</span><strong>{html.escape(value)}</strong></div>" for label, value in kpi_cards)
    insight_html = "\n".join(f"<li>{html.escape(item)}</li>" for item in insights)
    recommendation_html = "\n".join(f"<li>{html.escape(item)}</li>" for item in recommendations)

    content = f"""<!doctype html>
<html lang="en">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>Retail Sales Report</title>
  <style>
    :root {{ --ink:#111827; --muted:#64748b; --line:#dbe3ef; --blue:#1d4ed8; --green:#15803d; --orange:#ea580c; --bg:#f8fafc; }}
    body {{ margin:0; background:var(--bg); color:var(--ink); font-family:Segoe UI, Arial, sans-serif; line-height:1.5; }}
    main {{ max-width:1180px; margin:0 auto; padding:36px 24px 56px; }}
    header {{ border-bottom:1px solid var(--line); padding-bottom:20px; margin-bottom:24px; }}
    h1 {{ margin:0 0 8px; font-size:34px; }}
    h2 {{ margin-top:34px; font-size:22px; }}
    .meta {{ color:var(--muted); }}
    .kpis {{ display:grid; grid-template-columns:repeat(auto-fit,minmax(160px,1fr)); gap:14px; margin:22px 0; }}
    .kpis div {{ background:white; border:1px solid var(--line); border-top:4px solid var(--blue); padding:14px; border-radius:8px; }}
    .kpis span {{ display:block; color:var(--muted); font-size:13px; }}
    .kpis strong {{ display:block; margin-top:6px; font-size:22px; }}
    .grid {{ display:grid; grid-template-columns:repeat(auto-fit,minmax(380px,1fr)); gap:18px; }}
    figure {{ margin:0; background:white; border:1px solid var(--line); border-radius:8px; padding:12px; }}
    img {{ width:100%; height:auto; display:block; }}
    section {{ background:white; border:1px solid var(--line); border-radius:8px; padding:18px; margin-top:18px; }}
    li {{ margin:8px 0; }}
    .data-table {{ width:100%; border-collapse:collapse; font-size:14px; }}
    .data-table th {{ text-align:left; background:#eaf1fb; color:#0f172a; }}
    .data-table th, .data-table td {{ padding:9px 10px; border-bottom:1px solid var(--line); }}
    .data-table td:nth-child(n+2), .data-table th:nth-child(n+2) {{ text-align:right; }}
  </style>
</head>
<body>
<main>
  <header>
    <h1>Retail Sales Report</h1>
    <div class="meta">Source: {html.escape(input_path.name)} | Date range: {kpis['date_start'].date()} to {kpis['date_end'].date()} | Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}</div>
  </header>
  <div class="kpis">{kpi_html}</div>
  <section><h2>Insights</h2><ul>{insight_html}</ul></section>
  <section><h2>Recommendations</h2><ul>{recommendation_html}</ul></section>
  <h2>Visualizations</h2>
  <div class="grid">{chart_html}</div>
  <section><h2>Performance by Category</h2>{html_table(summaries['category'], symbol)}</section>
  <section><h2>Performance by Region</h2>{html_table(summaries['region'], symbol)}</section>
  <section><h2>Performance by Segment</h2>{html_table(summaries['segment'], symbol)}</section>
</main>
</body>
</html>"""
    output_path.write_text(content, encoding="utf-8")


def write_dataframe(ws, df: pd.DataFrame, table_name: str | None = None) -> None:
    for row in dataframe_to_rows(df, index=False, header=True):
        ws.append(list(row))

    if len(df) > 0:
        max_row = len(df) + 1
        max_col = len(df.columns)
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(max_col)}{max_row}"
        if table_name:
            table = Table(displayName=table_name, ref=f"A1:{get_column_letter(max_col)}{max_row}")
            table.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium2",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False,
            )
            ws.add_table(table)

    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.alignment = Alignment(horizontal="center")

    for column_cells in ws.columns:
        column_letter = get_column_letter(column_cells[0].column)
        header = str(column_cells[0].value)
        max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells[: min(len(column_cells), 50)])
        width = min(max(max_length + 2, len(header) + 2, 11), 28)
        ws.column_dimensions[column_letter].width = width

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            header = ws.cell(row=1, column=cell.column).value
            if header in {"Sales", "Profit"}:
                cell.number_format = '$#,##0'
            elif header in {"ProfitMargin", "SalesShare", "ProfitShare"}:
                cell.number_format = "0.0%"
            elif header == "Date":
                cell.number_format = "yyyy-mm-dd"


def style_dashboard(ws) -> None:
    ws.sheet_view.showGridLines = False
    for col in range(1, 11):
        ws.column_dimensions[get_column_letter(col)].width = 16
    for row in range(1, 40):
        ws.row_dimensions[row].height = 20


def add_kpi_card(ws, cell: str, label: str, value: object, note: str, fill: str) -> None:
    start = ws[cell]
    row = start.row
    col = start.column
    ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + 1)
    ws.merge_cells(start_row=row + 1, start_column=col, end_row=row + 1, end_column=col + 1)
    ws.merge_cells(start_row=row + 2, start_column=col, end_row=row + 2, end_column=col + 1)
    for r in range(row, row + 3):
        for c in range(col, col + 2):
            ws.cell(r, c).fill = PatternFill("solid", fgColor=fill)
            ws.cell(r, c).border = Border(
                left=Side(style="thin", color="D9E2F3"),
                right=Side(style="thin", color="D9E2F3"),
                top=Side(style="thin", color="D9E2F3"),
                bottom=Side(style="thin", color="D9E2F3"),
            )
    ws.cell(row, col).value = label
    ws.cell(row, col).font = Font(bold=True, color="334155", size=10)
    ws.cell(row + 1, col).value = value
    ws.cell(row + 1, col).font = Font(bold=True, color="0F172A", size=18)
    ws.cell(row + 2, col).value = note
    ws.cell(row + 2, col).font = Font(color="64748B", size=9)
    for r in range(row, row + 3):
        ws.cell(r, col).alignment = Alignment(horizontal="center", vertical="center")


def add_dashboard_charts(wb: Workbook) -> None:
    dash = wb["Dashboard"]
    daily = wb["Daily Trend"]
    category = wb["By Category"]
    region = wb["By Region"]
    segment = wb["By Segment"]

    daily_max = daily.max_row
    category_max = category.max_row
    region_max = region.max_row
    segment_max = segment.max_row

    line = LineChart()
    line.title = "Daily Sales and Profit"
    line.y_axis.title = "Amount"
    line.x_axis.title = "Date"
    line.height = 8
    line.width = 15
    data = Reference(daily, min_col=3, max_col=4, min_row=1, max_row=daily_max)
    cats = Reference(daily, min_col=1, min_row=2, max_row=daily_max)
    line.add_data(data, titles_from_data=True)
    line.set_categories(cats)
    line.y_axis.numFmt = '$#,##0'
    dash.add_chart(line, "A20")

    category_chart = BarChart()
    category_chart.type = "bar"
    category_chart.title = "Sales and Profit by Category"
    category_chart.y_axis.title = "Category"
    category_chart.x_axis.title = "Amount"
    category_chart.height = 8
    category_chart.width = 15
    data = Reference(category, min_col=3, max_col=4, min_row=1, max_row=category_max)
    cats = Reference(category, min_col=1, min_row=2, max_row=category_max)
    category_chart.add_data(data, titles_from_data=True)
    category_chart.set_categories(cats)
    category_chart.x_axis.numFmt = '$#,##0'
    dash.add_chart(category_chart, "F20")

    region_chart = BarChart()
    region_chart.type = "bar"
    region_chart.title = "Sales and Profit by Region"
    region_chart.y_axis.title = "Region"
    region_chart.x_axis.title = "Amount"
    region_chart.height = 8
    region_chart.width = 15
    data = Reference(region, min_col=3, max_col=4, min_row=1, max_row=region_max)
    cats = Reference(region, min_col=1, min_row=2, max_row=region_max)
    region_chart.add_data(data, titles_from_data=True)
    region_chart.set_categories(cats)
    region_chart.x_axis.numFmt = '$#,##0'
    dash.add_chart(region_chart, "A36")

    margin_chart = BarChart()
    margin_chart.type = "col"
    margin_chart.title = "Profit Margin by Segment"
    margin_chart.y_axis.title = "Profit Margin"
    margin_chart.x_axis.title = "Segment"
    margin_chart.height = 8
    margin_chart.width = 15
    data = Reference(segment, min_col=5, max_col=5, min_row=1, max_row=segment_max)
    cats = Reference(segment, min_col=1, min_row=2, max_row=segment_max)
    margin_chart.add_data(data, titles_from_data=True)
    margin_chart.set_categories(cats)
    margin_chart.y_axis.numFmt = "0.0%"
    dash.add_chart(margin_chart, "F36")


def write_workbook(
    workbook_path: Path,
    cleaned: pd.DataFrame,
    quality: pd.DataFrame,
    summaries: dict[str, pd.DataFrame],
    kpis: dict[str, object],
    insights: list[str],
    recommendations: list[str],
    symbol: str,
) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Dashboard"
    style_dashboard(ws)

    ws.merge_cells("A1:J1")
    ws["A1"] = "Retail Sales Dashboard"
    ws["A1"].font = Font(bold=True, size=22, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="1F4E78")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:J2")
    ws["A2"] = f"Source: retail_sales.csv | Date range: {kpis['date_start'].date()} to {kpis['date_end'].date()}"
    ws["A2"].font = Font(color="475569", italic=True)
    ws["A2"].alignment = Alignment(horizontal="center")

    kpi_fill = "EEF4FF"
    add_kpi_card(ws, "A4", "Orders", number(kpis["orders"]), "Unique order IDs", kpi_fill)
    add_kpi_card(ws, "C4", "Total Sales", money(kpis["total_sales"], symbol), "Gross sales", kpi_fill)
    add_kpi_card(ws, "E4", "Total Profit", money(kpis["total_profit"], symbol), "After cost proxy", kpi_fill)
    add_kpi_card(ws, "G4", "Profit Margin", percent(kpis["profit_margin"]), "Profit / Sales", kpi_fill)
    add_kpi_card(ws, "I4", "Avg Order Value", money(kpis["avg_order_value"], symbol), "Sales / orders", kpi_fill)

    ws.merge_cells("A9:J9")
    ws["A9"] = "Key Insights"
    ws["A9"].font = Font(bold=True, size=14, color="FFFFFF")
    ws["A9"].fill = PatternFill("solid", fgColor="15803D")
    for idx, insight in enumerate(insights, start=10):
        ws.merge_cells(start_row=idx, start_column=1, end_row=idx, end_column=10)
        ws.cell(idx, 1).value = f"- {insight}"
        ws.cell(idx, 1).alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[idx].height = 34

    start_reco = 10 + len(insights) + 1
    ws.merge_cells(start_row=start_reco, start_column=1, end_row=start_reco, end_column=10)
    ws.cell(start_reco, 1).value = "Recommended Actions"
    ws.cell(start_reco, 1).font = Font(bold=True, size=14, color="FFFFFF")
    ws.cell(start_reco, 1).fill = PatternFill("solid", fgColor="EA580C")
    for idx, item in enumerate(recommendations, start=start_reco + 1):
        ws.merge_cells(start_row=idx, start_column=1, end_row=idx, end_column=10)
        ws.cell(idx, 1).value = f"- {item}"
        ws.cell(idx, 1).alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[idx].height = 34

    sheets = [
        ("Daily Trend", summaries["daily"], "DailyTrendTable"),
        ("By Region", summaries["region"], "RegionSummaryTable"),
        ("By Category", summaries["category"], "CategorySummaryTable"),
        ("By Segment", summaries["segment"], "SegmentSummaryTable"),
        ("Cleaned Data", cleaned, "CleanedDataTable"),
        ("Data Quality", quality, "DataQualityTable"),
    ]
    for sheet_name, df, table_name in sheets:
        sheet = wb.create_sheet(sheet_name)
        write_dataframe(sheet, df, table_name)

    add_dashboard_charts(wb)
    wb.save(workbook_path)


def write_outputs(input_path: Path, output_dir: Path, currency_symbol: str) -> ReportOutputs:
    output_dir.mkdir(parents=True, exist_ok=True)
    cleaned, quality = load_and_clean(input_path)
    summaries = build_summaries(cleaned)
    kpis = build_kpis(cleaned, summaries)
    insights = build_insights(cleaned, summaries, kpis, currency_symbol)
    recommendations = build_recommendations(summaries)
    charts = create_charts(summaries, output_dir, currency_symbol)

    cleaned_csv = output_dir / "retail_sales_cleaned.csv"
    markdown_report = output_dir / "retail_sales_report.md"
    html_report = output_dir / "retail_sales_report.html"
    workbook = output_dir / "retail_sales_report.xlsx"

    cleaned.to_csv(cleaned_csv, index=False)
    write_markdown(markdown_report, input_path, kpis, insights, recommendations, summaries, charts, currency_symbol)
    write_html(html_report, input_path, kpis, insights, recommendations, summaries, charts, currency_symbol)
    write_workbook(workbook, cleaned, quality, summaries, kpis, insights, recommendations, currency_symbol)

    return ReportOutputs(cleaned_csv, markdown_report, html_report, workbook, charts)


def verify_outputs(outputs: ReportOutputs) -> None:
    required = [outputs.cleaned_csv, outputs.markdown_report, outputs.html_report, outputs.workbook, *outputs.charts]
    missing = [path for path in required if not path.exists() or path.stat().st_size == 0]
    if missing:
        raise FileNotFoundError(f"Missing or empty output files: {missing}")

    wb = load_workbook(outputs.workbook, data_only=False)
    expected_sheets = {"Dashboard", "Daily Trend", "By Region", "By Category", "By Segment", "Cleaned Data", "Data Quality"}
    missing_sheets = expected_sheets.difference(wb.sheetnames)
    if missing_sheets:
        raise ValueError(f"Workbook missing sheets: {sorted(missing_sheets)}")

    if len(wb["Dashboard"]._charts) < 4:
        raise ValueError("Dashboard should contain at least 4 native Excel charts.")


def main() -> None:
    args = parse_args()
    input_path = Path(args.input)
    output_dir = Path(args.output_dir)
    outputs = write_outputs(input_path, output_dir, args.currency_symbol)
    verify_outputs(outputs)

    print(f"Cleaned CSV: {outputs.cleaned_csv}")
    print(f"Markdown report: {outputs.markdown_report}")
    print(f"HTML report: {outputs.html_report}")
    print(f"Excel workbook: {outputs.workbook}")
    print("Charts:")
    for chart in outputs.charts:
        print(f"  - {chart}")


if __name__ == "__main__":
    main()
